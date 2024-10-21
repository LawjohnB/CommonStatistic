from openpyxl import load_workbook
import sqlite3
from datetime import datetime
import select_data_from_db as sdfdb
import msvcrt

start_date = datetime.strptime('08.10.2024', '%d.%m.%Y').date()
end_date = datetime.strptime('21.10.2024', '%d.%m.%Y').date()
start_year = datetime(datetime.now().year, 1, 1).date()


# номер текущей недели в соответствии с выбранной датой
current_week = datetime.isocalendar(end_date)[1]
# в названии таблицы присутствует номер недели и текущий год
current_year = datetime.now().year
# название таблицы включает в себя номер недели и года
table_prefix = '_'.join(map(str, (current_week, current_year)))


	# словарь столбцов (columns) типов отчёта: смещение столбца относительно номера недели
	# 2 полугодие
# recieved - поступило материалов в неделе
# mat_recieved - Количество уголовных дел, по которым поступили	(всех материалов)
# primary_recieved - поступило первичных
# additional_recieved - поступило дополнительных
# repeated_recieved - поступило повторных
# mats_complited - количество уголовных дел, по которым проведены (всего материалов?)




colls_shift = {'recieved': -15, 'mat_recieved': 14, 'primary_recieved': 43,
	'additional_recieved': 72, 'repeated_recieved': 101, 'comission_recieved': 130,
	'additional_comission_recieved': 159, 'repeated_comission_recieved': 188, 
	'complex_recieved': 217, 'additional_complex_recieved': 246, 'repeated_complex_recieved': 275,
	'mats_complited': 331, 'complited': 307, 'complited_objs': 334, 'without_exec': 363,
	'complited_terms_less_month': 392, 'complited_terms_1_month': 421, 'complited_terms_2_3_month': 450,
	'complited_terms_more_3_month': 479, 'primary_completed': 508, 'additional_completed': 537,
	'repeated_completed': 566, 'comission_completed': 595, 'additional_comission_completed': 624,
	'repeated_comission_completed': 653, 'complex_completed': 682, 'additional_complex_completed': 711,
	'repeated_complex_completed': 740, 'npv': 769, 'facts_est': 827}

stable_colls = {'recieved_total': 9, 'mat_recieved_total': 38, 'primary_recieved_total': 67,
	'additional_recieved_total': 96, 'repeated_recieved_total': 125, 'comission_recieved_total': 154,
	'additional_comission_recieved_total': 183, 'repeated_comission_recieved_total': 212,
	'complex_recieved_total': 241, 'additional_complex_recieved_total': 270, 'repeated_complex_recieved_total': 299,
	'mats_complited': 331, 'complited_exps_total': 328, 'complited_objs_total': 329,
	'without_exec_total': 387, 'complited_terms_less_month_total': 416, 'complited_terms_1_month_total': 445,
	'complited_terms_2_3_month_total': 474, 'complited_terms_more_3_month_total': 503,
	'primary_completed_total': 532, 'additional_completed_total': 561,
	'repeated_completed_total': 590, 'comission_completed_total': 619, 'additional_comission_completed_total': 648,
	'repeated_comission_completed_total': 677, 'complex_completed_total': 706,
	'additional_complex_completed_total': 735, 'repeated_complex_completed_total': 764,
	'npv_total': 793, 'persons_est_total': 822, 'facts_est_total': 851,
	'exps_in_work': 880, 'objs_in_work': 881, 'in_work_terms_less_month': 882,
	'in_work_terms_1_month': 883, 'in_work_terms_2_3_month': 884, 'in_work_terms_more_3_month': 885}

	# Загрузка книги для чтения и записи
	# data_only=True - считывать только значения (не формулы)
#print('Загрузка книги')

# wb = load_workbook('ЮФ_НЕД_42_II.xlsx', data_only=True)
# ws = wb["ЭКСПЕРТИЗЫ II полугодие 2024"]


# Заполнение строк 
# ИАЭ - строка 18
table_name = f'Week_{table_prefix}_ФОНО_Exps'
human_readable_table_name = table_name.split('_')[-2]


	## ПОСТУПИВШИЕ
# Поступило за неделю (всего)
recieved_on_week = sdfdb.get_recieved_exp_count(start_date, end_date, table_name)
# Поступило с начала года (всего)
recieved_total = sdfdb.get_recieved_exp_count(start_year, end_date, table_name)
# Поступило за неделю материалов (уголовных дел)
mat_recieved = sdfdb.get_recieved_mats_count(start_date, end_date, table_name)
# Поступило за материалов (уголовных дел/МП) (всего)
mat_recieved_total = sdfdb.get_recieved_mats_count(start_year, end_date, table_name)
# поступило первичных (обычных) за неделю
primary_recieved = sdfdb.get_primary_recieved_count(start_date, end_date, table_name)
# поступило первичных (обычных) (всего)
primary_recieved_total = sdfdb.get_primary_recieved_count(start_year, end_date, table_name)
# поступило дополнительных за неделю
additional_recieved = sdfdb.get_additional_recieved_count(start_date, end_date, table_name)
# поступило дополнительных всего
additional_recieved_total = sdfdb.get_additional_recieved_count(start_year, end_date, table_name)
# поступило повторных
repeated_recieved = sdfdb.get_repeated_recieved_count(start_date, end_date, table_name)
# поступило повторных всего
repeated_recieved_total = sdfdb.get_repeated_recieved_count(start_year, end_date, table_name)
# поступило комиссионных (первичных)
comission_recieved = sdfdb.get_comission_recieved_count(start_date, end_date, table_name)
# поступило комиссионных (первичных) всего 
comission_recieved_total = sdfdb.get_comission_recieved_count(start_year, end_date, table_name)
# поступило комиссионных дополнительных
additional_comission_recieved = sdfdb.get_additional_comission_recieved_count(start_date, end_date, table_name)
# поступило комиссионных дополнительных всего 
additional_comission_recieved_total = sdfdb.get_additional_comission_recieved_count(start_year, end_date, table_name)
# поступило дополнительных комиссионных
repeated_comission_recieved = sdfdb.get_repeated_comission_recieved_count(start_date, end_date, table_name)
# поступило дополнительных комиссионных всего 
repeated_comission_recieved_total = sdfdb.get_repeated_comission_recieved_count(start_year, end_date, table_name)
# поступило комплексных (первичных)
complex_recieved = sdfdb.get_complex_recieved_count(start_date, end_date, table_name)
# поступило комплексных (первичных) всего
complex_recieved_total = sdfdb.get_complex_recieved_count(start_year, end_date, table_name)
# поступило комплексных дополнительных
additional_complex_recieved = sdfdb.get_additional_complex_recieved_count(start_date, end_date, table_name)
# поступило комплексных дополнительных
additional_complex_recieved_total = sdfdb.get_additional_complex_recieved_count(start_year, end_date, table_name)
# поступило комплексных повторных
repeated_complex_recieved = sdfdb.get_repeated_complex_recieved_count(start_date, end_date, table_name)
# поступило комплексных повторных всего 
repeated_complex_recieved_total = sdfdb.get_repeated_complex_recieved_count(start_year, end_date, table_name)

# проверка совпадения суммы поступивших
if recieved_on_week != sum((primary_recieved, additional_recieved, repeated_recieved,
	comission_recieved, additional_comission_recieved, repeated_comission_recieved,
	complex_recieved, additional_complex_recieved, repeated_complex_recieved)):
	print(f'Внимание! Проверьте количество поступивших экспертиз ({human_readable_table_name})')


## ВЫПОЛНЕННЫЕ
# материалов по выполненным С НАЧАЛА ГОДА!
mats_complited = sdfdb.get_mats_complited(start_year, end_date, table_name)
# выполнено за неделю
complited_exps = sdfdb.get_complited(start_date, end_date, table_name)
# выполнено всего 
complited_exps_total = sdfdb.get_complited(start_year, end_date, table_name)
# объектов по выполненным
complited_objs = sdfdb.get_complited_objs(start_date, end_date, table_name)
# объектов по выполненным всего
complited_objs_total = sdfdb.get_complited_objs(start_year, end_date, table_name)


		# Для выполнения данных запросов требуется Python > 3.8
	# сроки исполненных
status = "exp_status = 'Сдана'"
# до 1 месяца
terms_condition = "month = '00'" # срок до 1 месяца
# за неделю
complited_terms_less_month = sdfdb.get_complited_terms(start_date, end_date, table_name, status, terms_condition)
# всего
complited_terms_less_month_total = sdfdb.get_complited_terms(start_year, end_date, table_name, status, terms_condition)

# от 1 до 2 месяцев (не включая 2)
terms_condition = "month = '01'" # срок 1 месяц
# за неделю
complited_terms_1_month = sdfdb.get_complited_terms(start_date, end_date, table_name, status, terms_condition)
# всего
complited_terms_1_month_total = sdfdb.get_complited_terms(start_year, end_date, table_name, status, terms_condition)

# от 2 до 3 месяцев (включительно)
terms_condition = "month = '02' or month = '03'" # срок 2 или 3 месяца
# за неделю
complited_terms_2_3_month = sdfdb.get_complited_terms(start_date, end_date, table_name, status, terms_condition)
# всего
complited_terms_2_3_month_total = sdfdb.get_complited_terms(start_year, end_date, table_name, status, terms_condition)

# свыше 3 месяцев
terms_condition = "month NOT IN ('00', '01', '02', '03')" # срок более 3 месяцев (количество месяцев - строка)
# за неделю
complited_terms_more_3_month = sdfdb.get_complited_terms(start_date, end_date, table_name, status, terms_condition)
# всего
complited_terms_more_3_month_total = sdfdb.get_complited_terms(start_year, end_date, table_name, status, terms_condition)

# проверка количество выполненных
if complited_exps != sum((complited_terms_less_month, complited_terms_1_month,
					complited_terms_2_3_month, complited_terms_more_3_month)):
	print(f'Внимание! Проверьте количество выполненных экспертиз и сроки исполнения ({human_readable_table_name})')

	# возвращено без исполнения
without_exec = sdfdb.get_without_exec(start_date, end_date, table_name)
without_exec_total = sdfdb.get_without_exec(start_year, end_date, table_name)

# выполненные по видам
# выполнено первичных (обычных) за неделю
primary_completed = sdfdb.get_primary_completed_count(start_date, end_date, table_name)
# выполнено первичных (обычных) (всего)
primary_completed_total = sdfdb.get_primary_completed_count(start_year, end_date, table_name)
# выполнено дополнительных за неделю
additional_completed = sdfdb.get_additional_completed_count(start_date, end_date, table_name)
# выполнено дополнительных всего
additional_completed_total = sdfdb.get_additional_completed_count(start_year, end_date, table_name)
# выполнено повторных
repeated_completed = sdfdb.get_repeated_completed_count(start_date, end_date, table_name)
# выполнено повторных всего
repeated_completed_total = sdfdb.get_repeated_completed_count(start_year, end_date, table_name)
# выполнено комиссионных (первичных)
comission_completed = sdfdb.get_comission_completed_count(start_date, end_date, table_name)
# выполнено комиссионных (первичных) всего 
comission_completed_total = sdfdb.get_comission_completed_count(start_year, end_date, table_name)
# выполнено комиссионных дополнительных
additional_comission_completed = sdfdb.get_additional_comission_completed_count(start_date, end_date, table_name)
# выполнено комиссионных дополнительных всего 
additional_comission_completed_total = sdfdb.get_additional_comission_completed_count(start_year, end_date, table_name)
# выполнено дополнительных комиссионных
repeated_comission_completed = sdfdb.get_repeated_comission_completed_count(start_date, end_date, table_name)
# выполнено дополнительных комиссионных всего 
repeated_comission_completed_total = sdfdb.get_repeated_comission_completed_count(start_year, end_date, table_name)
# выполнено комплексных (первичных)
complex_completed = sdfdb.get_complex_completed_count(start_date, end_date, table_name)
# выполнено комплексных (первичных) всего
complex_completed_total = sdfdb.get_complex_completed_count(start_year, end_date, table_name)
# выполнено комплексных дополнительных
additional_complex_completed = sdfdb.get_additional_complex_completed_count(start_date, end_date, table_name)
# выполнено комплексных дополнительных
additional_complex_completed_total = sdfdb.get_additional_complex_completed_count(start_year, end_date, table_name)
# выполнено комплексных повторных
repeated_complex_completed = sdfdb.get_repeated_complex_completed_count(start_date, end_date, table_name)
# выполнено комплексных повторных всего 
repeated_complex_completed_total = sdfdb.get_repeated_complex_completed_count(start_year, end_date, table_name)

# проверка совпадения суммы выполненных
if complited_exps != sum((primary_completed, additional_completed, repeated_completed,
    comission_completed, additional_comission_completed, repeated_comission_completed,
    complex_completed, additional_complex_completed, repeated_complex_completed)):
	print(f'Внимание! Проверьте количество выполненных экспертиз и их типы ({human_readable_table_name})')

# НПВ
# за неделю
npv = sdfdb.get_npv_count(start_date, end_date, table_name)
# всего
npv_total = sdfdb.get_npv_count(start_year, end_date, table_name)

# установлено лиц
persons_est = sdfdb.get_persons_est_count(start_date, end_date, table_name)
persons_est_total = sdfdb.get_persons_est_count(start_year, end_date, table_name)

# установлено факторов
facts_est = sdfdb.get_facts_est_count(start_date, end_date, table_name)
facts_est_total = sdfdb.get_facts_est_count(start_year, end_date, table_name)

	# В ПРОИЗВОДСТВЕ
# всего экспертиз
exps_in_work = sdfdb.get_work_count(table_name)
# всего объектов
objs_in_work = sdfdb.get_work_objs_count(table_name)

# сроки по находящимся в работе
# end_date - дата отчёта, на которую производится расчёт
status = "exp_status = 'В производстве'"
# до 1 месяца
terms_condition = "month = '00'" # срок до 1 месяца
in_work_terms_less_month = sdfdb.get_work_terms(end_date, table_name, status, terms_condition)
# от 1 до 2 месяцев (не включая 2)
terms_condition = "month = '01'" # срок 1 месяц
in_work_terms_1_month = sdfdb.get_work_terms(end_date, table_name, status, terms_condition)
# от 2 до 3 месяцев (включительно)
terms_condition = "month = '02' or month = '03'" # срок 2 или 3 месяца
in_work_terms_2_3_month = sdfdb.get_work_terms(end_date, table_name, status, terms_condition)
# свыше 3 месяцев
terms_condition = "month NOT IN ('00', '01', '02', '03')" # срок более 3 месяцев (количество месяцев - строка)
in_work_terms_more_3_month = sdfdb.get_work_terms(end_date, table_name, status, terms_condition)

# проверка количество выполненных
if exps_in_work != sum((in_work_terms_less_month, in_work_terms_1_month, in_work_terms_2_3_month,
				in_work_terms_more_3_month)):
	print(f'Внимание! Проверьте количество выполненных экспертиз и сроки исполнения ({human_readable_table_name})')


# print("Press any key to continue...")
# msvcrt.getch()
	

# for row in temp_rows:
# 	print(f'Заполнение строки {temp_rows[row]}')
# 	coll = current_week - 15
# 	ws.cell(row=row, column=coll).value = f'{temp_rows[row]} поступило'
# 	coll = current_week + 14
# 	ws.cell(row=row, column=coll).value = f'{temp_rows[row]} УД поступило'
# 	coll = current_week + 43
# 	ws.cell(row=row, column=coll).value = f'{temp_rows[row]} первичных поступило'

# print('Сохранение данных')
# wb.save("TEST.xlsx") 