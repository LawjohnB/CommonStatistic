import os
import glob
import xlrd
import xlwt
from openpyxl import Workbook

# функция проверки наличия правильного количества файлов нужного вида
def check_files_count(files):
    if len(files) != 5:
        print(f'Количество файлов {files} некорректно!')


# обход каталогов для построения всех файлов Excel (файлы регионов)
xlsx_files = glob.glob('./Журналы регионов/*/*.xls*')


# распределение файлов Excel по типам
xlsx_exps = [file for file in xlsx_files if 'Журнал экспертиз' in file]
xlsx_issls = [file for file in xlsx_files if 'Журнал исследований' in file]
xlsx_sipd =[file for file in xlsx_files if 'Журнал следственных действий' in file]
xlsx_consults = [file for file in xlsx_files if 'Журнал консультаций' in file]

# контрольная проверка наличия всех файлов (по 5 от каждого региона)
check_files_count(xlsx_exps)
check_files_count(xlsx_issls)
check_files_count(xlsx_sipd)
check_files_count(xlsx_consults)

sheets = {}

# Цикл по файлам экспертиз
for excel_file in xlsx_exps:
    exps = xlrd.open_workbook(excel_file, on_demand = True)
    for sheet_name in exps.sheet_names():
        if sheet_name in ('Отдел', 'Недельный отчёт по Экспертам', 'Экспертизы',\
                        'Месячный отчёт по Экспертам', 'Инструкции', 'Справочник'):
            continue
        current_sheet = exps.sheet_by_name(sheet_name)
        rows = [current_sheet.row_values(x) for x in range(1, current_sheet.nrows)]
        sheets[sheet_name] = sheets.get(sheet_name, []) + rows
    exps.release_resources()
    del exps

# # Создание нового файла
# xlsx_join_exps = xlwt.Workbook()

# for sheet_name, values in sheets.items():
#     sheet = xlsx_join_exps.add_sheet(sheet_name)
#     sheet.write(0, 0, values) 

# xlsx_join_exps.save("Журнал экспертиз (общий).xlsx")
wb = Workbook()
for sheet_name, values in sheets.items():
    ws = wb.create_sheet(sheet_name)
    ws.title = sheet_name
    for v in values:
        if not v[0]:
            continue
        ws.append(v)

wb.save('Общий журнал.xlsx')